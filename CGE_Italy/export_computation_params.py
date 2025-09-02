#!/usr/bin/env python3
"""
Export Core Mathematical Parameters
Focus on actual coefficients and parameters used in CGE equations
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

from calibrate import model_data


def export_computation_parameters():
    """Export the mathematical parameters used in CGE computations"""

    # Load data
    sam_path = 'data/SAM.xlsx'
    sam = pd.read_excel(sam_path, index_col=0)

    h = ['Labour', 'Capital']
    ind = ['Agriculture', 'Industry', 'Electricity', 'Gas', 'Other Energy',
           'Road Transport', 'Rail Transport', 'Air Transport', 'Water Transport',
           'Other Transport', 'other Sectors (14)']

    data = model_data(sam, h, ind)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79",
                              end_color="1F4E79", fill_type="solid")

    # 1. Input-Output Matrix (aij coefficients)
    ws_io_matrix = wb.create_sheet("Input-Output Matrix (aij)")

    # Create the full I-O coefficient matrix
    io_matrix = pd.DataFrame(index=ind, columns=ind, dtype=float)
    for i in ind:
        for j in ind:
            if data.Z0[j] > 0:
                io_matrix.loc[i, j] = data.X0.loc[i, j] / data.Z0[j]
            else:
                io_matrix.loc[i, j] = 0.0

    # Add to worksheet
    ws_io_matrix.append(["FROM \\ TO"] + list(ind))
    for i, sector_i in enumerate(ind):
        row_data = [sector_i] + \
            [f"{io_matrix.loc[sector_i, sector_j]:.6f}" for sector_j in ind]
        ws_io_matrix.append(row_data)

    # Style header
    for cell in ws_io_matrix[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_io_matrix.column_dimensions['A'].width = 15
    for i in range(len(ind)):
        col_letter = chr(66 + i)  # B, C, D, ...
        ws_io_matrix.column_dimensions[col_letter].width = 10

    # 2. Factor Share Matrix (αij)
    ws_factor_shares = wb.create_sheet("Factor Shares (alpha_ij)")

    factor_shares = pd.DataFrame(index=h, columns=ind, dtype=float)
    for factor in h:
        for sector in ind:
            total_factors = data.F0.loc['Labour',
                                        sector] + data.F0.loc['Capital', sector]
            if total_factors > 0:
                factor_shares.loc[factor,
                                  sector] = data.F0.loc[factor, sector] / total_factors
            else:
                factor_shares.loc[factor, sector] = 0.5  # Default equal shares

    # Add to worksheet
    ws_factor_shares.append(["FACTOR \\ SECTOR"] + list(ind))
    for factor in h:
        row_data = [factor] + \
            [f"{factor_shares.loc[factor, sector]:.4f}" for sector in ind]
        ws_factor_shares.append(row_data)

    # Add verification row (should sum to 1.0)
    verification = ["SUM (should = 1.0)"] + \
        [f"{factor_shares[sector].sum():.4f}" for sector in ind]
    ws_factor_shares.append(verification)

    # Style header
    for cell in ws_factor_shares[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_factor_shares.column_dimensions['A'].width = 18
    for i in range(len(ind)):
        col_letter = chr(66 + i)
        ws_factor_shares.column_dimensions[col_letter].width = 12

    # 3. Elasticity Parameters Table
    ws_elasticities = wb.create_sheet("Elasticity Parameters")

    # Define elasticity parameters used in equations
    elasticity_data = [
        ["Parameter", "Symbol", "Value", "Sectors", "Mathematical Use"],
        ["CES Production Elasticity", "σ", "0.8", "All",
            "Y = A[αL*L^ρ + αK*K^ρ]^(1/ρ), ρ=(σ-1)/σ"],
        ["Armington Trade Elasticity", "σA", "2.0",
            "All", "Q = [δD*QD^ρ + δM*QM^ρ]^(1/ρ)"],
        ["Own Price Elasticity", "ηii", "-0.5",
            "All", "∂ln(Qi)/∂ln(Pi) = ηii"],
        ["Income Elasticity", "ηY", "1.0", "All", "∂ln(Qi)/∂ln(Y) = ηY"],
        ["Cross Price Elasticity", "ηij", "0.2",
            "All", "∂ln(Qi)/∂ln(Pj) = ηij"],
        ["", "", "", "", ""],
        ["SECTOR-SPECIFIC ELASTICITIES", "", "", "", ""],
        ["Energy Price Elasticity", "ηE", "-0.8",
            "Energy", "Electricity, Gas, Other Energy"],
        ["Transport Price Elasticity", "ηT", "-0.6",
            "Transport", "Road, Rail, Air, Water Transport"],
        ["Industry Price Elasticity", "ηI", "-0.7",
            "Industry", "Manufacturing sector"],
        ["Carbon Price Elasticity", "ηC", "-0.3",
            "ETS", "Response to carbon pricing"],
        ["", "", "", "", ""],
        ["TRADE ELASTICITIES", "", "", "", ""],
        ["Export Demand Elasticity", "ηX", "-1.5",
            "All", "Foreign demand for Italian exports"],
        ["Import Supply Elasticity", "ηM", "1.8", "All",
            "Foreign supply of imports to Italy"],
        ["Terms of Trade Elasticity", "ηTOT", "0.5", "All", "Price transmission"],
    ]

    for row in elasticity_data:
        ws_elasticities.append(row)

    # Style header
    for cell in ws_elasticities[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_elasticities.column_dimensions['A'].width = 25
    ws_elasticities.column_dimensions['B'].width = 8
    ws_elasticities.column_dimensions['C'].width = 10
    ws_elasticities.column_dimensions['D'].width = 15
    ws_elasticities.column_dimensions['E'].width = 45

    # 4. Carbon Coefficients Matrix
    ws_carbon_coeff = wb.create_sheet("Carbon Coefficients")

    # Carbon intensity by sector
    carbon_intensities = {
        'Agriculture': 0.120, 'Industry': 0.280, 'Electricity': 0.350,
        'Gas': 0.400, 'Other Energy': 0.420, 'Road Transport': 0.180,
        'Rail Transport': 0.045, 'Air Transport': 0.920, 'Water Transport': 0.650,
        'Other Transport': 0.150, 'other Sectors (14)': 0.080
    }

    carbon_data = [
        ["Sector", "Carbon Intensity (ci)", "Unit",
         "ETS Coverage", "Mathematical Use"],
    ]

    for sector in ind:
        intensity = carbon_intensities.get(sector, 0.1)
        ets_coverage = "ETS1" if sector in ['Electricity', 'Industry', 'Other Energy'] else \
            "ETS2 (2027+)" if 'Transport' in sector else "Not covered"
        carbon_data.append([
            sector, f"{intensity:.3f}", "tCO2/Million EUR", ets_coverage,
            f"Ei = ci * Zi * (1-γ)^t"
        ])

    for row in carbon_data:
        ws_carbon_coeff.append(row)

    # Style header
    for cell in ws_carbon_coeff[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_carbon_coeff.column_dimensions['A'].width = 18
    ws_carbon_coeff.column_dimensions['B'].width = 18
    ws_carbon_coeff.column_dimensions['C'].width = 18
    ws_carbon_coeff.column_dimensions['D'].width = 15
    ws_carbon_coeff.column_dimensions['E'].width = 25

    # 5. Dynamic Coefficients
    ws_dynamic_coeff = wb.create_sheet("Dynamic Coefficients")

    dynamic_data = [
        ["Parameter", "Symbol", "Value", "Equation", "Mathematical Role"],
        ["Depreciation Rate", "δ", "0.05",
            "K(t+1) = (1-δ)K(t) + I(t)", "Capital stock evolution"],
        ["TFP Growth Rate", "gA", "0.015",
            "A(t) = A(0)(1+gA)^t", "Productivity growth"],
        ["Population Growth", "n", "0.001",
            "POP(t) = POP(0)(1+n)^t", "Demographic dynamics"],
        ["Labor Growth", "gL", "0.002",
            "L(t) = L(0)(1+gL)^t", "Labor force growth"],
        ["Savings Rate", "s", "0.20", "S = s * Y", "Savings function"],
        ["Investment Adj. Speed", "φ", "2.0",
            "I = I(-1) + φ(K* - K(-1))", "Capital adjustment"],
        ["", "", "", "", ""],
        ["CARBON PARAMETERS", "", "", "", ""],
        ["Emissions Decline", "γ", "0.02",
            "Ei(t) = ci*Zi(t)*(1-γ)^t", "Technology improvement"],
        ["Carbon Price Growth", "gc", "0.05",
            "Pc(t) = Pc(0)*(1+gc)^t", "ETS price escalation"],
        ["Carbon Leakage", "λ", "0.15",
            "ΔE_foreign = λ * ΔE_domestic", "Production shifting"],
        ["", "", "", "", ""],
        ["ADJUSTMENT SPEEDS", "", "", "", ""],
        ["Price Adjustment", "θp", "0.5",
            "P(t) = P(t-1) + θp(P* - P(t-1))", "Price dynamics"],
        ["Quantity Adjustment", "θq", "0.3",
            "Q(t) = Q(t-1) + θq(Q* - Q(t-1))", "Quantity dynamics"],
        ["Wage Adjustment", "θw", "0.4",
            "W(t) = W(t-1) + θw(W* - W(t-1))", "Labor market dynamics"],
    ]

    for row in dynamic_data:
        ws_dynamic_coeff.append(row)

    # Style header
    for cell in ws_dynamic_coeff[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_dynamic_coeff.column_dimensions['A'].width = 20
    ws_dynamic_coeff.column_dimensions['B'].width = 8
    ws_dynamic_coeff.column_dimensions['C'].width = 10
    ws_dynamic_coeff.column_dimensions['D'].width = 35
    ws_dynamic_coeff.column_dimensions['E'].width = 30

    # 6. CES Function Parameters
    ws_ces = wb.create_sheet("CES Function Parameters")

    # Calculate CES parameters
    sigma = 0.8
    rho = (sigma - 1) / sigma

    ces_data = [
        ["CES Parameter", "Symbol", "Value", "Mathematical Definition"],
        ["Elasticity of Substitution", "σ", f"{sigma:.3f}", "σ = 1/(1-ρ)"],
        ["CES Parameter", "ρ", f"{rho:.6f}", "ρ = (σ-1)/σ"],
        ["", "", "", ""],
        ["SECTOR-SPECIFIC CES PARAMETERS", "", "", ""],
    ]

    for sector in ind:
        total_factors = data.F0.loc['Labour',
                                    sector] + data.F0.loc['Capital', sector]
        if total_factors > 0:
            alpha_L = data.F0.loc['Labour', sector] / total_factors
            alpha_K = data.F0.loc['Capital', sector] / total_factors

            ces_data.extend([
                [f"Labor Share - {sector}", f"αL_{sector[:8]}",
                    f"{alpha_L:.6f}", f"Labor weight in CES function"],
                [f"Capital Share - {sector}", f"αK_{sector[:8]}",
                    f"{alpha_K:.6f}", f"Capital weight in CES function"],
            ])

    # Add CES equation
    ces_data.extend([
        ["", "", "", ""],
        ["CES PRODUCTION FUNCTION", "", "", ""],
        ["General Form", "Y",
            "A[αL*L^ρ + αK*K^ρ]^(1/ρ)", "Constant Elasticity of Substitution"],
        ["With σ=0.8", "Y",
            f"A[αL*L^{rho:.3f} + αK*K^{rho:.3f}]^{1/rho:.3f}", "Actual CES function used"],
    ])

    for row in ces_data:
        ws_ces.append(row)

    # Style header
    for cell in ws_ces[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_ces.column_dimensions['A'].width = 30
    ws_ces.column_dimensions['B'].width = 15
    ws_ces.column_dimensions['C'].width = 12
    ws_ces.column_dimensions['D'].width = 40

    # Save file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"results/cge_computation_parameters_{timestamp}.xlsx"

    wb.save(filename)

    print("="*70)
    print("MATHEMATICAL COMPUTATION PARAMETERS EXPORTED")
    print("="*70)
    print(f"File: {filename}")
    print()
    print("📋 EXCEL SHEETS CREATED:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    print()
    print("🔢 KEY MATHEMATICAL PARAMETERS INCLUDED:")
    print(
        f"  • Input-Output Coefficients: {len(ind)}×{len(ind)} = {len(ind)**2} coefficients")
    print(
        f"  • Factor Share Parameters: {len(h)}×{len(ind)} = {len(h)*len(ind)} shares")
    print(
        f"  • Elasticity Parameters: {len(ind)} sectors × 4 elasticity types")
    print(f"  • Carbon Coefficients: {len(ind)} emission intensities")
    print(f"  • Dynamic Parameters: 13 growth and adjustment rates")
    print(f"  • CES Parameters: {len(ind)*2} production function weights")
    print()
    print("📐 MATHEMATICAL EQUATIONS COVERED:")
    print("  • Production: Yi = Ai[αLi*Li^ρ + αKi*Ki^ρ]^(1/ρ)")
    print("  • Intermediate Demand: Xij = aij * Zj")
    print("  • Final Demand: Qi = fi(Pi, Y, βi)")
    print("  • Emissions: Ei = ci * Zi * (1-γ)^t")
    print("  • Capital Dynamics: K(t+1) = (1-δ)K(t) + I(t)")
    print("="*70)

    return filename


if __name__ == "__main__":
    export_computation_parameters()
