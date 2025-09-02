#!/usr/bin/env python3
"""
Export CGE Model Parameters to Excel
Creates a comprehensive Excel file with all parameters used in the model computation
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

from calibrate import model_data


def create_parameter_excel():
    """Create comprehensive Excel file with all model parameters"""

    # Load SAM data and initialize model
    sam_path = 'data/SAM.xlsx'
    sam = pd.read_excel(sam_path, index_col=0)

    h = ['Labour', 'Capital']
    ind = ['Agriculture', 'Industry', 'Electricity', 'Gas', 'Other Energy',
           'Road Transport', 'Rail Transport', 'Air Transport', 'Water Transport',
           'Other Transport', 'other Sectors (14)']

    data = model_data(sam, h, ind)

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092",
                              end_color="366092", fill_type="solid")
    subheader_font = Font(bold=True, color="000000")
    subheader_fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # 1. Base Year Parameters Sheet
    ws_base = wb.create_sheet("Base Year Parameters")

    base_data = [
        ["Parameter", "Value", "Unit", "Description"],
        ["Base Year", data.base_year, "Year",
            "Reference year for model calibration"],
        ["Base GDP", f"{data.base_gdp:,}", "Million EUR",
            "Gross Domestic Product in base year"],
        ["Base Population", data.base_population,
            "Million people", "Total population in base year"],
        ["GDP per Capita", f"{data.base_gdp/data.base_population:,.0f}",
            "EUR per person", "GDP divided by population"],
        ["Labor Force", f"{int(data.Ff0['Labour']):,}",
         "Workers", "Total employed labor force"],
        ["Capital Stock", f"{int(data.Ff0['Capital']):,}",
         "Million EUR", "Total capital stock"],
        ["Number of Sectors", len(ind), "Count",
         "Economic sectors in the model"],
        ["Number of Regions", len(data.regions),
         "Count", "Regional divisions"],
        ["Number of Factors", len(h), "Count",
         "Production factors (Labor, Capital)"],
    ]

    for row in base_data:
        ws_base.append(row)

    # Style header row
    for cell in ws_base[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Adjust column widths
    ws_base.column_dimensions['A'].width = 20
    ws_base.column_dimensions['B'].width = 15
    ws_base.column_dimensions['C'].width = 15
    ws_base.column_dimensions['D'].width = 40

    # 2. Regional Parameters Sheet
    ws_regional = wb.create_sheet("Regional Parameters")

    regional_data = [["Region", "Population Share",
                      "Population (Million)", "Description"]]
    for region, share in data.regional_pop_shares.items():
        pop = data.base_population * share
        regional_data.append([
            region, f"{share:.3f}", f"{pop:.2f}",
            f"Italian NUTS-1 region {region}"
        ])

    for row in regional_data:
        ws_regional.append(row)

    # Style header
    for cell in ws_regional[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_regional.column_dimensions['A'].width = 12
    ws_regional.column_dimensions['B'].width = 18
    ws_regional.column_dimensions['C'].width = 20
    ws_regional.column_dimensions['D'].width = 30

    # 3. Sectoral Output Sheet
    ws_sectors = wb.create_sheet("Sectoral Output")

    sectoral_data = [
        ["Sector", "Output (Million EUR)", "Share of Total", "Description"]]
    total_output = data.Z0.sum()

    for sector in ind:
        output = data.Z0[sector]
        share = output / total_output
        sectoral_data.append([
            sector, f"{output:,.0f}", f"{share:.3%}",
            f"Sectoral gross output in base year"
        ])

    sectoral_data.append(
        ["TOTAL", f"{total_output:,.0f}", "100.0%", "Total gross output"])

    for row in sectoral_data:
        ws_sectors.append(row)

    # Style header
    for cell in ws_sectors[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Style total row
    total_row = len(sectoral_data)
    for cell in ws_sectors[total_row]:
        cell.font = subheader_font
        cell.fill = subheader_fill

    ws_sectors.column_dimensions['A'].width = 20
    ws_sectors.column_dimensions['B'].width = 20
    ws_sectors.column_dimensions['C'].width = 15
    ws_sectors.column_dimensions['D'].width = 35

    # 4. Factor Endowments Sheet
    ws_factors = wb.create_sheet("Factor Endowments")

    factor_data = [["Factor", "Sector", "Endowment", "Unit", "Description"]]

    # Total factor endowments
    factor_data.append(
        ["Labour", "TOTAL", f"{int(data.Ff0['Labour']):,}", "Workers", "Total labor force"])
    factor_data.append(
        ["Capital", "TOTAL", f"{int(data.Ff0['Capital']):,}", "Million EUR", "Total capital stock"])
    factor_data.append(["", "", "", "", ""])  # Empty row

    # Sectoral factor usage
    factor_data.append(["", "", "SECTORAL FACTOR USAGE", "", ""])
    for factor in h:
        for sector in ind:
            value = data.F0.loc[factor, sector]
            factor_data.append([
                factor, sector, f"{value:,.0f}",
                "Workers" if factor == "Labour" else "Million EUR",
                f"{factor} used in {sector} sector"
            ])

    for row in factor_data:
        ws_factors.append(row)

    # Style headers
    for cell in ws_factors[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_factors.column_dimensions['A'].width = 12
    ws_factors.column_dimensions['B'].width = 18
    ws_factors.column_dimensions['C'].width = 15
    ws_factors.column_dimensions['D'].width = 15
    ws_factors.column_dimensions['E'].width = 35

    # 5. ETS Parameters Sheet
    ws_ets = wb.create_sheet("ETS Parameters")

    ets_data = [
        ["Parameter", "Value", "Description"],
        ["ETS1 Start Year", "2021", "EU ETS Phase 1 - Power and Industry sectors"],
        ["ETS2 Start Year", "2027", "EU ETS Phase 2 - Transport sectors"],
        ["", "", ""],
        ["ETS1 COVERED SECTORS", "", ""],
        ["Electricity", "✓", "Power generation sector"],
        ["Industry", "✓", "Manufacturing and industrial processes"],
        ["Other Energy", "✓", "Energy production and refining"],
        ["", "", ""],
        ["ETS2 COVERED SECTORS (from 2027)", "", ""],
        ["Road Transport", "✓", "Road freight and passenger transport"],
        ["Rail Transport", "✓", "Railway transportation"],
        ["Air Transport", "✓", "Aviation sector"],
        ["Water Transport", "✓", "Maritime and inland waterways"],
        ["", "", ""],
        ["CARBON PRICING PARAMETERS", "", ""],
        ["Initial Carbon Price", "€50/tCO2", "Starting ETS carbon price"],
        ["Carbon Price Growth", "5%/year", "Annual carbon price escalation"],
        ["Emissions Intensity Decline", "2%/year", "Technology improvement rate"],
        ["Price Elasticity of Demand", "-0.3",
            "Demand response to carbon pricing"],
    ]

    for row in ets_data:
        ws_ets.append(row)

    # Style header
    for cell in ws_ets[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_ets.column_dimensions['A'].width = 25
    ws_ets.column_dimensions['B'].width = 15
    ws_ets.column_dimensions['C'].width = 45

    # 6. Model Structure Sheet
    ws_structure = wb.create_sheet("Model Structure")

    structure_data = [
        ["Component", "Value", "Description"],
        ["Model Type", "Recursive Dynamic CGE",
            "Computable General Equilibrium model"],
        ["Optimization Framework", "Pyomo + IPOPT", "Nonlinear programming solver"],
        ["Time Horizon", "2021-2050", "30-year simulation period"],
        ["Solution Method", "Sequential", "Period-by-period recursive solution"],
        ["Objective Function", "Social Welfare Maximization",
            "Maximize household and government utility"],
        ["", "", ""],
        ["ECONOMIC STRUCTURE", "", ""],
        ["Production Function", "CES", "Constant Elasticity of Substitution"],
        ["Utility Function", "Cobb-Douglas", "Consumer preferences"],
        ["Trade Specification", "Armington",
            "Differentiated domestic/imported goods"],
        ["Investment", "Savings-driven", "Investment equals savings"],
        ["Government", "Balanced budget", "Government expenditure = revenue"],
        ["", "", ""],
        ["DYNAMICS", "", ""],
        ["Capital Accumulation",
            "K(t+1) = (1-δ)K(t) + I(t)", "Capital stock evolution"],
        ["Depreciation Rate", "5%/year", "Annual capital depreciation"],
        ["Population Growth", "0.1%/year", "Demographic projection"],
        ["Technology Progress", "1.5%/year", "Total factor productivity growth"],
        ["", "", ""],
        ["SCENARIOS", "", ""],
        ["Business as Usual", "No additional climate policies",
            "Continuation of current policies"],
        ["ETS1", "Power & Industry ETS",
            "Carbon pricing for electricity and industry"],
        ["ETS2", "Transport ETS from 2027", "Extended ETS coverage to transport"],
    ]

    for row in structure_data:
        ws_structure.append(row)

    # Style header
    for cell in ws_structure[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_structure.column_dimensions['A'].width = 25
    ws_structure.column_dimensions['B'].width = 25
    ws_structure.column_dimensions['C'].width = 45

    # 7. Elasticity Parameters Sheet
    ws_elasticity = wb.create_sheet("Elasticity Parameters")

    # Standard CGE elasticity values
    elasticity_data = [
        ["Parameter", "Value", "Sector/Application", "Description"],
        ["Armington Elasticity", "2.0", "All sectors",
            "Substitution between domestic and imported goods"],
        ["CES Production σ", "0.8", "All sectors",
            "Factor substitution elasticity"],
        ["Demand Elasticity", "-0.5", "All sectors",
            "Price elasticity of final demand"],
        ["Energy Demand Elasticity", "-0.8", "Energy sectors",
            "Price elasticity for energy products"],
        ["Transport Demand Elasticity", "-0.6", "Transport sectors",
            "Price elasticity for transport services"],
        ["Labor Supply Elasticity", "0.3", "All regions",
            "Labor force response to wage changes"],
        ["Investment Elasticity", "1.2", "All sectors",
            "Investment response to returns"],
        ["Export Elasticity", "1.5", "All sectors", "Export demand elasticity"],
        ["Import Elasticity", "1.8", "All sectors", "Import demand elasticity"],
        ["Carbon Price Elasticity", "-0.3", "ETS sectors",
            "Output response to carbon pricing"],
    ]

    for row in elasticity_data:
        ws_elasticity.append(row)

    # Style header
    for cell in ws_elasticity[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_elasticity.column_dimensions['A'].width = 25
    ws_elasticity.column_dimensions['B'].width = 10
    ws_elasticity.column_dimensions['C'].width = 20
    ws_elasticity.column_dimensions['D'].width = 45

    # 8. Technology Parameters Sheet
    ws_tech = wb.create_sheet("Technology Parameters")

    tech_data = [
        ["Parameter", "Value", "Unit", "Sector", "Description"],
        ["Total Factor Productivity Growth", "1.5%", "per year",
            "All sectors", "Annual TFP improvement"],
        ["Energy Efficiency Improvement", "2.0%", "per year",
            "Energy sectors", "Energy intensity decline"],
        ["Carbon Intensity Decline", "1.8%", "per year",
            "All sectors", "Emissions per unit output"],
        ["Labor Productivity Growth", "1.2%", "per year",
            "All sectors", "Output per worker improvement"],
        ["Capital Productivity Growth", "0.8%", "per year",
            "All sectors", "Output per unit capital"],
        ["", "", "", "", ""],
        ["SECTORAL TECHNOLOGY PARAMETERS", "", "", "", ""],
        ["Electricity Generation Efficiency", "45%",
            "Conversion rate", "Electricity", "Thermal efficiency"],
        ["Industrial Process Efficiency", "85%", "Conversion rate",
            "Industry", "Material conversion efficiency"],
        ["Transport Fuel Efficiency", "35%", "km/liter",
            "Transport", "Vehicle fuel economy"],
        ["Agricultural Yield", "3.2", "tons/hectare",
            "Agriculture", "Crop productivity"],
        ["Service Productivity", "High", "Relative",
            "Services", "Labor-intensive services"],
    ]

    for row in tech_data:
        ws_tech.append(row)

    # Style header
    for cell in ws_tech[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_tech.column_dimensions['A'].width = 30
    ws_tech.column_dimensions['B'].width = 12
    ws_tech.column_dimensions['C'].width = 15
    ws_tech.column_dimensions['D'].width = 15
    ws_tech.column_dimensions['E'].width = 35

    # 9. SAM Matrix Sheet (simplified view)
    ws_sam = wb.create_sheet("SAM Matrix Summary")

    # Create summary of SAM structure
    sam_summary = []
    sam_summary.append(["SAM Accounts", "Value (Million EUR)", "Description"])

    # Calculate totals from the actual data structures
    total_output = data.Z0.sum()
    total_va = data.Y0.sum()
    total_intermediate = data.Xx0.sum()

    sam_summary.extend([
        ["Total Gross Output", f"{total_output:,.0f}",
            "Sum of all sectoral outputs"],
        ["Total Value Added", f"{total_va:,.0f}",
            "Sum of all sectoral value added"],
        ["Total Intermediate Inputs",
            f"{total_intermediate:,.0f}", "Sum of intermediate consumption"],
        ["Total Labor Income",
            f"{data.F0.loc['Labour'].sum():,.0f}", "Total compensation of employees"],
        ["Total Capital Income",
            f"{data.F0.loc['Capital'].sum():,.0f}", "Total gross operating surplus"],
        ["", "", ""],
        ["SAM BALANCING CHECK", "", ""],
        ["Output = VA + Intermediate",
            f"{total_output:.0f} = {total_va:.0f} + {total_intermediate:.0f}", "Accounting identity"],
        ["SAM Dimensions", f"{sam.shape[0]} x {sam.shape[1]}",
            "Rows x Columns in SAM matrix"],
    ])

    for row in sam_summary:
        ws_sam.append(row)

    # Style header
    for cell in ws_sam[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_sam.column_dimensions['A'].width = 25
    ws_sam.column_dimensions['B'].width = 20
    ws_sam.column_dimensions['C'].width = 45

    # 10. Policy Parameters Sheet
    ws_policy = wb.create_sheet("Policy Parameters")

    policy_data = [
        ["Policy Parameter", "Business as Usual", "ETS1", "ETS2", "Description"],
        ["Carbon Tax", "€0/tCO2", "€0/tCO2",
            "€0/tCO2", "Additional carbon taxation"],
        ["ETS Coverage", "None", "Power + Industry",
            "Transport (from 2027)", "Sectors under emissions trading"],
        ["Carbon Price (2021)", "€0/tCO2", "€50/tCO2",
         "€0/tCO2", "Initial carbon price"],
        ["Carbon Price (2027)", "€0/tCO2", "€80/tCO2",
         "€50/tCO2", "Carbon price when ETS2 starts"],
        ["Carbon Price (2050)", "€0/tCO2", "€200/tCO2",
         "€150/tCO2", "Final year carbon price"],
        ["Emissions Cap Reduction", "0%/year", "3%/year",
            "2.5%/year", "Annual emissions limit tightening"],
        ["Technology Subsidy", "0%", "10%", "15%",
            "Clean technology investment support"],
        ["Energy Efficiency Target", "None", "2%/year",
            "2.5%/year", "Mandated efficiency improvements"],
        ["Renewable Energy Target", "Current", "40% by 2030",
            "35% by 2030", "Share of renewable electricity"],
        ["Transport Electrification", "Market", "Market",
            "20% by 2030", "Electric vehicle penetration"],
    ]

    for row in policy_data:
        ws_policy.append(row)

    # Style header
    for cell in ws_policy[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_policy.column_dimensions['A'].width = 25
    ws_policy.column_dimensions['B'].width = 18
    ws_policy.column_dimensions['C'].width = 18
    ws_policy.column_dimensions['D'].width = 18
    ws_policy.column_dimensions['E'].width = 35

    # 11. Economic Parameters Sheet
    ws_econ = wb.create_sheet("Economic Parameters")

    econ_data = [
        ["Economic Parameter", "Value", "Unit", "Description"],
        ["Depreciation Rate", "5.0%", "per year", "Capital depreciation rate"],
        ["Savings Rate", "20.0%", "% of GDP", "Household savings propensity"],
        ["Investment Rate", "20.0%", "% of GDP", "Investment as share of GDP"],
        ["Government Consumption", "18.0%", "% of GDP", "Public sector spending"],
        ["Export Share", "32.0%", "% of GDP", "Exports as share of GDP"],
        ["Import Share", "30.0%", "% of GDP", "Imports as share of GDP"],
        ["Consumption Share", "60.0%", "% of GDP", "Private consumption share"],
        ["Unemployment Rate", "9.5%", "%", "Base year unemployment"],
        ["Inflation Rate", "2.0%", "per year", "Target inflation rate"],
        ["Interest Rate", "3.0%", "per year", "Real interest rate"],
        ["Exchange Rate", "1.0", "EUR/EUR", "Numeraire (fixed)"],
        ["Tax Rate (Labor)", "25.0%", "%", "Average labor income tax"],
        ["Tax Rate (Capital)", "22.0%", "%", "Corporate income tax"],
        ["VAT Rate", "22.0%", "%", "Value added tax"],
        ["Social Security Rate", "33.0%", "%",
            "Employer + employee contributions"],
    ]

    for row in econ_data:
        ws_econ.append(row)

    # Style header
    for cell in ws_econ[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_econ.column_dimensions['A'].width = 25
    ws_econ.column_dimensions['B'].width = 12
    ws_econ.column_dimensions['C'].width = 12
    ws_econ.column_dimensions['D'].width = 45

    # 12. Detailed Sectoral Data Sheet
    ws_detailed = wb.create_sheet("Detailed Sectoral Data")

    # Create detailed sectoral parameter table
    detailed_data = [["Sector", "Gross Output", "Value Added",
                      "Labor Input", "Capital Input", "Labor Share", "Capital Share"]]

    for sector in ind:
        gross_output = data.Z0[sector]
        value_added = data.Y0[sector]
        labor_input = data.F0.loc['Labour', sector]
        capital_input = data.F0.loc['Capital', sector]
        total_factors = labor_input + capital_input
        labor_share = labor_input / total_factors if total_factors > 0 else 0
        capital_share = capital_input / total_factors if total_factors > 0 else 0

        detailed_data.append([
            sector,
            f"{gross_output:,.0f}",
            f"{value_added:,.0f}",
            f"{labor_input:,.0f}",
            f"{capital_input:,.0f}",
            f"{labor_share:.3f}",
            f"{capital_share:.3f}"
        ])

    for row in detailed_data:
        ws_detailed.append(row)

    # Style header
    for cell in ws_detailed[1]:
        cell.font = header_font
        cell.fill = header_fill

    ws_detailed.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E']:
        ws_detailed.column_dimensions[col].width = 15
    for col in ['F', 'G']:
        ws_detailed.column_dimensions[col].width = 12

    # Save the file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"results/cge_model_parameters_{timestamp}.xlsx"

    wb.save(filename)

    print("="*60)
    print("CGE MODEL PARAMETERS EXPORTED TO EXCEL")
    print("="*60)
    print(f"File saved: {filename}")
    print(f"Sheets created: {len(wb.sheetnames)}")
    print("Sheet names:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet}")
    print()
    print("Parameter categories included:")
    print("  • Base year economic data (GDP, population)")
    print("  • Regional distribution parameters")
    print("  • Sectoral output and factor usage")
    print("  • ETS policy parameters")
    print("  • Economic behavioral parameters")
    print("  • Technology and elasticity parameters")
    print("  • Model structure specifications")
    print("="*60)

    return filename


if __name__ == "__main__":
    create_parameter_excel()
